import time
import streamlit as st
import pandas as pd
import re
import unicodedata
import os
import pickle
import tempfile
from docx import Document
from docx.document import Document as _Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from langchain_community.vectorstores import FAISS
from langchain_community.embeddings.fastembed import FastEmbedEmbeddings
from langchain.docstore.document import Document as LangchainDocument
from difflib import SequenceMatcher
from rapidfuzz import fuzz, process
from langchain_groq import ChatGroq
import tiktoken
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Document Comparison Tool", layout="wide")

if 'comparison_completed' not in st.session_state:
    st.session_state.comparison_completed = False
if 'company_faiss' not in st.session_state:
    st.session_state.company_faiss = None
if 'customer_faiss' not in st.session_state:
    st.session_state.customer_faiss = None
if 'final_df' not in st.session_state:
    st.session_state.final_df = None
if 'output_df' not in st.session_state:
    st.session_state.output_df = None
if 'processing_step' not in st.session_state:
    st.session_state.processing_step = None
if 'temp_dir' not in st.session_state:
    st.session_state.temp_dir = tempfile.mkdtemp()

temp_dir = st.session_state.temp_dir

def preprocess_text(text):
    if not text:
        return ""
    text = re.sub(r"<.*?>", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = unicodedata.normalize("NFKD", text)
    text = text.lower()
    text = re.sub(r"[^\w\s.]", "", text)
    return text

def count_tokens(text, encoding_name="cl100k_base"):
    encoding = tiktoken.get_encoding(encoding_name)
    return len(encoding.encode(text))

def truncate_text(text, max_tokens=100):
    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(text)
    return encoding.decode(tokens[:max_tokens])

def iter_block_items(parent):
    if isinstance(parent, _Document):
        parent_elm = parent.element.body
    else:
        parent_elm = parent._element
    for child in parent_elm.iterchildren():
        if child.tag.endswith("p"):
            yield Paragraph(child, parent)
        elif child.tag.endswith("tbl"):
            yield Table(child, parent)

def extract_document_content(docx_path):
    """Extract all content from the document without relying on markers"""
    extracted_data = []
    doc = Document(docx_path)
    order_counter = 0

    for block in iter_block_items(doc):
        if block.__class__.__name__ == "Paragraph":
            para_text = block.text.strip()
            if para_text:
                extracted_data.append({
                    "text": para_text,
                    "type": "paragraph",
                    "order": order_counter
                })
                order_counter += 1
        elif block.__class__.__name__ == "Table":
            table_data = []
            for row in block.rows:
                row_cells = [cell.text.strip() for cell in row.cells]
                table_data.append(" | ".join(row_cells))
            table_text = "\n".join(table_data)
            if table_text:
                extracted_data.append({
                    "text": f"[TABLE] {table_text}",
                    "type": "table",
                    "order": order_counter
                })
                order_counter += 1
    return extracted_data

@st.cache_resource
def process_document(docx_path):
    """Process document and prepare for comparison"""
    progress_bar = st.progress(0)
    extracted_data = extract_document_content(docx_path)
    
    # Combine all document content into one text block
    full_text = "\n".join([item["text"] for item in extracted_data])
    
    # Create a single document object for the entire content
    doc_obj = LangchainDocument(
        page_content=full_text,
        metadata={"source": os.path.basename(docx_path)}
    )
    
    # Store in FAISS for retrieval
    embedding_model = FastEmbedEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
    faiss_index = FAISS.from_documents([doc_obj], embedding_model)
    faiss_index.documents = [doc_obj]  # Store original document
    
    progress_bar.empty()
    return faiss_index, extracted_data

def extract_customer_number(filename):
    if '__' in filename:
        number_part = filename.split('__')[0]
        return number_part
    return os.path.splitext(filename)[0]

def create_direct_comparison_prompt(company_content, customer_content, customer_number="All Samples"):
    prompt = (
        "You are a document comparison expert. You will analyze differences between company (Filed Copy) and customer document sections "
        "and produce a consolidated report. The format of your response is critical.\n\n"

        "IMPORTANT: Do not miss out on any differences as per the input documents\n"
        
        "IMPORTANT: Create a table with EXACTLY these column headers:\n"
        "1. 'Samples affected' - Should always be '{customer_number}'\n"
        "2. 'Observation - Category' - Must be one of these two categories:\n"
        "   - 'Mismatch of content between Filed Copy and customer copy': Use ONLY when similar content exists in BOTH documents but with differences\n"
        "   - 'Available in Filed Copy but missing in Customer Copy': Use ONLY when content exists in Filed copy but is COMPLETELY ABSENT from customer copy\n"
        "3. 'Page' - The specific document section(s) EXACTLY affected as per the below:\n"
        "   - 'Preamble'\n"
        "   - 'Terms and conditions'\n"
        "   - 'Annexure 1'\n"
        "   - 'Annexure AA'\n"
        "   - 'Forwarding Letter'\n"
        "   - 'Ombudsman Page'\n"
        "   - 'Schedule'\n"
        "4. 'Sub-category of Observation' - Comprehensive summary of ALL differences for that page\n\n"
        "5. For each page, provide a COMPREHENSIVE summary of ALL differences in the 'Sub-category of Observation' column.\n"
        "   IMPORTANT: Use bullet points (•) instead of numbers for listing differences in the 'Sub-category of Observation' column.\n"
        "6. Group all pages with 'Mismatch of content between Filed Copy and customer copy' together first, "
        "   then group all pages with 'Available in Filed Copy but missing in Customer Copy'.\n\n"
        
        "CLEAR DISTINCTION BETWEEN CATEGORIES:\n"
        "- 'Mismatch of content': Use ONLY when the same type of information exists in both documents but differs in details. Example: Company copy mentions '30-day period' while filed copy shows '15-day period'.\n"
        "- 'Available in Filed Copy but missing': Use ONLY when information appears in company document but is completely absent from customer document. Example: Company document has a section on 'Cancellation Policy' that doesn't appear at all in customer document.\n\n"
        
        "Here are the documents to compare:\n\n"
        
        f"COMPANY COPY (FILED COPY):\n{company_content}\n\n"
        f"CUSTOMER COPY:\n{customer_content}\n\n"
        "---\n\n"
    )
    
    prompt += (
        "INSTRUCTIONS:\n"
        "1. Compare the content of both documents carefully.\n"
        "2. Identify meaningful differences - ignore minor formatting, spacing, or punctuation differences.\n"
        "3. First determine if content is missing entirely (use 'Available in Filed Copy but missing') OR if similar content exists with differences (use 'Mismatch of content').\n"
        "4. Create ONE ROW PER PAGE in the output table, grouping by category.\n"
        "5. For each page, provide a COMPREHENSIVE summary of ALL differences in the 'Sub-category of Observation' column.\n"
        "6. Group all pages with 'Mismatch of content between Filed Copy and customer copy' together first, "
        "   then group all pages with 'Available in Filed Copy but missing in Customer Copy'.\n\n"
        
        "CRITICAL CATEGORIZATION RULES:\n"
        "- If similar content/information exists in both copies but with differences -> 'Mismatch of content'\n"
        "- If content/information exists in Filed copy but is completely absent in Customer copy -> 'Available in Filed Copy but missing'\n"
        "- Example of 'Mismatch': Both documents mention grace period but one says 30 days and other says 15 days\n"
        "- Example of 'Available but missing': Filed copy has contact details section that doesn't exist at all in customer copy\n\n"
        
        "Return your analysis as a JSON array where each object has these keys:\n"
        "- 'Samples affected': Always set to '{customer_number}'\n"
        "- 'Observation - Category': One of the two categories listed above\n"
        "- 'Page': One of the page categories mentioned above\n"
        "- 'Sub-category of Observation': Comprehensive summary of ALL differences for that page\n\n"
        
        "Example structure:\n"
        "[{\n"
        "  \"Samples affected\": \"{customer_number}\",\n"
        "  \"Observation - Category\": \"Mismatch of content between Filed Copy and customer copy\",\n"
        "  \"Page\": \"Forwarding Letter\",\n"
        "  \"Sub-category of Observation\": \"• Company copy mentions a 30-day response period while filed copy states 15 days.\n• Company copy includes different phone number than filed copy.\"\n"
        "},\n"
        "{\n"
        "  \"Samples affected\": \"{customer_number}\",\n"
        "  \"Observation - Category\": \"Available in Filed Copy but missing in Customer Copy\",\n"
        "  \"Page\": \"Terms and conditions\",\n"
        "  \"Sub-category of Observation\": \"• Section 4.2 on cancellation policy present in filed copy is completely missing from customer copy.\n• Appendix B with fee schedule present in filed copy doesn't exist in customer copy.\"\n"
        "}]\n\n"
        
        "IMPORTANT: Create exactly ONE row for each affected page within each category (do not create multiple rows for the same page in the same category). "
        "Make sure each 'Page' appears only once under each 'Observation - Category'. "
        "Group rows by 'Observation - Category'."
    )
    
    return prompt

def chunk_text_by_token_count(text, max_tokens=6000):
    """Split text into chunks based on token count limit"""
    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(text)
    
    chunks = []
    current_chunk = []
    current_count = 0
    
    for token in tokens:
        if current_count + 1 > max_tokens:
            chunks.append(encoding.decode(current_chunk))
            current_chunk = [token]
            current_count = 1
        else:
            current_chunk.append(token)
            current_count += 1
    
    if current_chunk:
        chunks.append(encoding.decode(current_chunk))
    
    return chunks

def organize_comparison_results(output_df):
    if output_df.empty:
        return output_df
        
    category_order = [
        "Mismatch of content between Filed Copy and customer copy",
        "Available in Filed Copy but missing in Customer Copy"
    ]
    
    page_order = [
        "Forwarding Letter", 
        "Schedule", 
        "Terms and conditions",
        "Ombudsman Page",
        "Annexure AA",
        "Annexure 1",
        "Preamble"
    ]
    
    output_df['category_order'] = output_df['Observation - Category'].apply(
        lambda x: category_order.index(x) if x in category_order else 999
    )
    
    output_df['page_order'] = output_df['Page'].apply(
        lambda x: page_order.index(x) if x in page_order else 999
    )
    
    sorted_df = output_df.sort_values(['category_order', 'page_order'])
    
    sorted_df = sorted_df.drop(['category_order', 'page_order'], axis=1)
    
    return sorted_df

@st.cache_resource
def direct_document_comparison(company_content, customer_content, groq_api_key, customer_number="All Samples"):
    # Truncate content to avoid token limit issues
    company_content_truncated = truncate_text(company_content, 4000)
    customer_content_truncated = truncate_text(customer_content, 4000)
    
    prompt = create_direct_comparison_prompt(company_content_truncated, customer_content_truncated, customer_number)
    
    groq_llm = ChatGroq(
        api_key=groq_api_key,
        model_name="llama3-8b-8192",
        max_tokens=2000
    )
    
    with st.spinner("Analyzing document differences..."):
        try:
            response = groq_llm.invoke([{"role": "user", "content": prompt}]).content
        except Exception as e:
            st.error(f"LLM Processing Error: {e}")
            return pd.DataFrame(columns=["Samples affected", "Observation - Category", "Page", "Sub-category of Observation"])
    
    if not response.strip():
        st.warning("Empty response from LLM.")
        return pd.DataFrame(columns=["Samples affected", "Observation - Category", "Page", "Sub-category of Observation"])
    
    try:
        response_text = response.strip()
        import re
        json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if json_match:
            json_text = json_match.group(0)
            results = json.loads(json_text)
        else:
            try:
                results = json.loads(response_text)
            except:
                import ast
                try:
                    results = ast.literal_eval(response_text)
                except:
                    st.error("Could not parse response as JSON")
                    st.code(response)
                    return pd.DataFrame(columns=["Samples affected", "Observation - Category", "Page", "Sub-category of Observation"])
        
        if results:
            for result in results:
                for key in list(result.keys()):
                    if key.lower() == "samples affected" or key.lower() == "samples_affected":
                        result["Samples affected"] = result.pop(key)
                    elif key.lower() == "observation - category" or key.lower() == "observation_category":
                        result["Observation - Category"] = result.pop(key)
                    elif key.lower() == "page":
                        result["Page"] = result.pop(key)
                    elif key.lower() == "sub-category of observation" or key.lower() == "sub_category":
                        result["Sub-category of Observation"] = result.pop(key)
        
            output_df = pd.DataFrame(results)
            
            required_columns = ["Samples affected", "Observation - Category", "Page", "Sub-category of Observation"]
            for col in required_columns:
                if col not in output_df.columns:
                    output_df[col] = ""
            
            output_df = output_df[required_columns]
            output_df = organize_comparison_results(output_df)
            
            return output_df
        else:
            return pd.DataFrame(columns=["Samples affected", "Observation - Category", "Page", "Sub-category of Observation"])
    except Exception as e:
        st.error(f"Error processing LLM response: {e}")
        st.code(response)
        return pd.DataFrame(columns=["Samples affected", "Observation - Category", "Page", "Sub-category of Observation"])

def save_uploaded_file(uploaded_file):
    if uploaded_file is not None:
        file_path = os.path.join(temp_dir, uploaded_file.name)
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return file_path
    return None

def run_direct_comparison(company_path, customer_path, groq_api_key):
    progress_container = st.empty()
    output_df = None
    
    try:
        customer_filename = os.path.basename(customer_path)
        customer_number = extract_customer_number(customer_filename)
        
        with progress_container.container():
            st.subheader("Step 1: Processing Company Document")
            st.text("Extracting company document content...")
            st.session_state.company_faiss, company_data = process_document(company_path)
            company_content = "\n".join([item["text"] for item in company_data])
            st.session_state.processing_step = "company_processed"
            st.success("✅ Company document processed successfully")

        with progress_container.container():
            st.subheader("Step 2: Processing Customer Document")
            st.text("Extracting customer document content...")
            st.session_state.customer_faiss, customer_data = process_document(customer_path)
            customer_content = "\n".join([item["text"] for item in customer_data])
            st.session_state.processing_step = "customer_processed"
            st.success("✅ Customer document processed successfully")

        with progress_container.container():
            st.subheader("Step 3: Analyzing Document Differences")
            output_df = direct_document_comparison(company_content, customer_content, groq_api_key, customer_number)
            
            if not output_df.empty:
                category_order = [
                    "Mismatch of content between Filed Copy and customer copy",
                    "Available in Filed Copy but missing in Customer Copy"
                ]
                
                for category in category_order:
                    if category not in output_df["Observation - Category"].values:
                        placeholder_row = {
                            "Samples affected": customer_number,
                            "Observation - Category": category,
                            "Page": "",
                            "Sub-category of Observation": ""
                        }
                        placeholder_df = pd.DataFrame([placeholder_row])
                        output_df = pd.concat([output_df, placeholder_df], ignore_index=True)
                
                output_df = output_df[output_df["Page"] != ""]
            
            output_df = organize_comparison_results(output_df)
            
            st.session_state.output_df = output_df
            
            if not output_df.empty:
                diff_count = len(output_df)
                st.session_state.final_df = pd.DataFrame({
                    "Section": output_df["Page"].tolist(),
                    "Comparison": ["DIFFERENT"] * diff_count,
                    "Difference": output_df["Sub-category of Observation"].tolist()
                })
            else:
                st.session_state.final_df = pd.DataFrame(columns=["Section", "Comparison", "Difference"])
                
            st.session_state.processing_step = "complete"
            st.session_state.comparison_completed = True
            st.success("✅ Document analysis completed")

        progress_container.empty()
        return True
        
    except Exception as e:
        if output_df is not None:
            st.session_state.output_df = output_df
        
        progress_container.error(f"Error occurred during processing: {e}")
        import traceback
        st.code(traceback.format_exc())
        return False

def display_results():
    st.header("Document Comparison Results")
    
    if st.session_state.comparison_completed:
        if st.button("Start New Comparison"):
            st.session_state.comparison_completed = False
            st.session_state.final_df = None
            st.session_state.output_df = None
            st.rerun()
    
    if st.session_state.final_df is not None:
        df_final = st.session_state.final_df
        
        st.subheader("Comparison Report")
        if st.session_state.output_df is not None and not st.session_state.output_df.empty:
            st.markdown("""
            <style>
            .comparison-table {
                width: 100%;
                text-align: left;
                border-collapse: collapse;
            }
            .comparison-table th {
                background-color: #87CEEB;
                color: black;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #ddd;
            }
            .comparison-table td {
                padding: 8px;
                border: 1px solid #ddd;
            }
            .comparison-table tr:nth-child(even) {
                background-color: #f9f9f9;
            }
            </style>
            """, unsafe_allow_html=True)
            
            html_table = st.session_state.output_df.to_html(
                classes="comparison-table", 
                index=False,
                escape=False
            )
            st.markdown(html_table, unsafe_allow_html=True)
            
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
            
            buffer = BytesIO()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Comparison Report"
            
            headers = ["Samples affected", "Observation - Category", "Page", "Sub-category of Observation"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            
            df = st.session_state.output_df
            current_category = None
            category_start_row = 2
            row_idx = 2
            
            for category in df['Observation - Category'].unique():
                category_df = df[df['Observation - Category'] == category]
                category_start_row = row_idx
                
                for _, row_data in category_df.iterrows():
                    cell = ws.cell(row=row_idx, column=1)
                    cell.value = row_data["Samples affected"]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    cell = ws.cell(row=row_idx, column=2)
                    cell.value = category if row_idx == category_start_row else None
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    cell = ws.cell(row=row_idx, column=3)
                    cell.value = row_data["Page"]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    cell = ws.cell(row=row_idx, column=4)
                    cell.value = row_data["Sub-category of Observation"]
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    row_idx += 1
                
                if row_idx > category_start_row + 1:
                    ws.merge_cells(start_row=category_start_row, start_column=2, end_row=row_idx-1, end_column=2)
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 60
            
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(
                label="Download Comparison Report",
                data=buffer,
                file_name="document_comparison_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("No differences found between the documents.")
    else:
        st.info("No comparison results available yet. Please run a comparison first.")

def main():
    st.title("Document Comparison Tool")
    
    with st.expander("How to Use"):
        st.markdown("""
        ### Instructions
        1. Upload the company document (DOCX format)
        2. Upload the customer document (DOCX format) 
        3. Enter your Groq API key
        4. Click "Run Comparison"
        
        ### About the Tool
        This tool compares two documents and uses AI to identify differences between them.
        It categorizes the differences for easy review.
        """)
    
    col1, col2 = st.columns(2)
    with col1:
        company_file = st.file_uploader("Upload Company Document (DOCX)", type=["docx"])
    with col2:
        customer_file = st.file_uploader("Upload Customer Document (DOCX)", type=["docx"])
    
    groq_api_key = st.text_input("Enter Groq API Key", type="password")
    
    if st.button("Run Comparison", disabled=not (company_file and customer_file and groq_api_key)):
        with st.spinner("Processing..."):
            company_path = save_uploaded_file(company_file)
            customer_path = save_uploaded_file(customer_file)
            
            success = run_direct_comparison(company_path, customer_path, groq_api_key)
            
            if success:
                st.success("Comparison completed successfully!")
            else:
                st.error("Comparison failed. Check logs for details.")
    
    if st.session_state.comparison_completed:
        display_results()

if __name__ == "__main__":
    main()
